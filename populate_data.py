#!/usr/bin/env python3
"""
Run this script to clear the database, and then populate it with data
"""
FILE_PATH = "api-utils/fittrack_data.xlsx"

# DELETE EXISTING DATABASE
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.sql import text

# Define the connection string
username = 'root'
password = 'password'
host = 'localhost'

# Create the engine to connect to the MySQL server
engine = create_engine(f'mysql+mysqldb://{username}:{password}@{host}', 
                      pool_pre_ping=True)

# Connect to the engine
connection = engine.connect()
SQL_STATEMENTS = [
    "DROP DATABASE IF EXISTS fittrack_dev_db;",
    "CREATE DATABASE IF NOT EXISTS fittrack_dev_db;",
    "CREATE DATABASE IF NOT EXISTS fittrack_test_db;",
    "CREATE USER IF NOT EXISTS 'fittrack_dev'@'localhost';",
    "SET PASSWORD FOR 'fittrack_dev'@'localhost' = 'fittrack_dev_pwd';",
    "GRANT ALL ON fittrack_dev_db.* TO 'fittrack_dev'@'localhost';",
    "GRANT ALL ON fittrack_test_db.* TO 'fittrack_dev'@'localhost';",
    "GRANT SELECT ON performance_schema.* TO 'fittrack_dev'@'localhost';",
    "FLUSH PRIVILEGES;",
]

try:
    # Execute SQL statements
    for stmt in SQL_STATEMENTS:
        connection.execute(text(stmt))

except SQLAlchemyError as e:
    print(f"Error: {e}")

finally:
    # Close the connection
    connection.close()
    print("DATABASE reset successfully.")

# POPULATE DATABASE
import openpyxl as op
from models.user import User, user_goals
from models.program import Program, program_goals
from models.routine import Routine
from models.workout_day import WorkoutDay
from models.article import Article
from models.goal import Goal
from models.body_focus import BodyFocus
from models.workout import Workout, workout_likes
from models.video import Video
from models.review import ProgramReview, WorkoutReview
from models import storage
from sqlalchemy import insert
from werkzeug.security import generate_password_hash
from datetime import datetime


wb = op.load_workbook(FILE_PATH, data_only=True)
# print(wb.worksheets)

entries = 0

def get_value(ws, row, col):
    return ws.cell(row, col).value

data_log = dict()

def worksheet(ws_name, obj, types=dict()):
    global entries
    ws = wb[ws_name]
    # map columns
    col_map = dict()
    for col in range(1, ws.max_column+1, 1):
        v = ws.cell(1, col).value
        if not v or len(v) == 0:
            break
        col_map[v] = col

    for row in range(2, ws.max_row+1, 1):
        attr_dict = dict()
        if ws.cell(row, 1).value is None or len(str(ws.cell(row, 1).value)) == 0:
            break;
        for attr, col in col_map.items():
            if attr in ['updated_at', 'created_at']:
                continue # skip
            if attr == 'password':
                value = generate_password_hash(get_value(ws, row, col),
                                               method="pbkdf2:sha1:1000",
                                               salt_length=8)
            elif attr in ['date', 'start_date' 'end_date']:
                value = datetime.strptime(get_value(ws, row, col), '%Y-%m-%d').date()
            elif attr in types:
                value = types[attr](get_value(ws, row, col))
            else:
                value = str(get_value(ws, row, col))
            attr_dict[attr] = value

        instance = obj(**attr_dict)
        instance.save()
        entries += 1
        obj_name = obj._sa_class_manager.class_.__name__
        data_log[f"{obj_name}.{instance.id}"] = instance

data_constraints = {
    'users': (User,
              {'weight': float, 'age': int}),
    'goals': (Goal, ),
    'body_focus': (BodyFocus,),
    'programs': (Program,
                 {'duration': int, 'difficulty': int}),
    'routines': (Routine,),
    'videos': (Video,
               {'thumbnail_width': int, 'thumbnail_height': int}),
    'workouts': (Workout,
                 {'workout_day': int}),
    'articles': (Article,),
    'workout_days': (WorkoutDay,
                     {'completed_status': int}),
    'program_reviews':  (ProgramReview,
                 {'rating': int}),
}


for ws_name, data in data_constraints.items():
    class_ = data[0]
    if len(data) > 1:
        types = data[1]
    else:
        types = dict()
    worksheet(ws_name, class_, types)


# populate association tables

association_tables = {
    'user_goals': user_goals,
    'program_goals': program_goals,
    'workout_likes': workout_likes,
}

for table, table_obj in association_tables.items():
    ws = wb[table]
    # map columns
    col_map = dict()
    for col in range(1, ws.max_column+1, 1):
        v = ws.cell(1, col).value
        if not v or len(v) == 0:
            break
        col_map[v] = col
    values_list = list()
    # loop through worksheet
    for row in range(2, ws.max_row+1, 1):
        attr_dict = dict()
        for attr, col in col_map.items():
            value = ws.cell(row, col).value
            attr_dict[attr] = value
        # add values to list
        values_list.append(attr_dict)
        entries += 1

    storage.close()
    # insert into table
    with storage._DBStorage__engine.connect() as conn:
        result = conn.execute(insert(table_obj), values_list,)
        conn.commit()

print('DONE!\tEntries: ', entries)
# print(data_log)
























